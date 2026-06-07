"""Raw sheet viewer endpoints — read V1/V2 as-is + edited V2.1 .xlsx download."""
import io
from pathlib import Path

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.config import settings
from app.deps import get_repo
from app.main import app

ROOT = Path(__file__).resolve().parents[2]
pytestmark = pytest.mark.skipif(
    not ((ROOT / "v1.xlsx").exists() and (ROOT / "v2.1.xlsx").exists()),
    reason="sample workbooks not present",
)


@pytest.fixture
def client(tmp_path, monkeypatch):
    monkeypatch.setattr(settings, "V1_PATH", str(ROOT / "v1.xlsx"))
    monkeypatch.setattr(settings, "V2_PATH", str(ROOT / "v2.1.xlsx"))
    monkeypatch.setattr(settings, "SNAPSHOT_PATH", str(tmp_path / "s.sqlite"))
    get_repo.cache_clear()
    with TestClient(app) as c:
        yield c
    get_repo.cache_clear()


def test_sheet_v1_as_is(client):
    g = client.get("/api/sheets/v1").json()
    assert "IS Reference Number" in g["columns"]
    assert "Level 1" in g["columns"]
    assert len(g["rows"]) > 0
    # every row carries its source Excel row index for targeted edits
    assert all("__row" in r for r in g["rows"])


def test_sheet_v2_as_is(client):
    g = client.get("/api/sheets/v2").json()
    assert "CC_V1_Mapping Entity" in g["columns"]
    # raw casing preserved (not normalised)
    joined = " ".join(r.get("Schema Name + JSON Path", "") for r in g["rows"])
    assert "ENtity" in joined


def test_fetch_entire_v1_row(client):
    grid = client.get("/api/sheets/v1").json()
    row_no = grid["rows"][1]["__row"]
    r = client.get(f"/api/sheets/v1/row/{row_no}")
    assert r.status_code == 200
    body = r.json()
    assert body["row"]["__row"] == row_no
    assert body["columns"] == grid["columns"]
    assert body["row"] == grid["rows"][1]          # same as the full-grid row


def test_fetch_entire_v2_row(client):
    grid = client.get("/api/sheets/v2").json()
    row_no = grid["rows"][0]["__row"]
    r = client.get(f"/api/sheets/v2/row/{row_no}")
    assert r.status_code == 200
    assert "CC_V1_Mapping Entity" in r.json()["columns"]


def test_fetch_missing_row_is_404(client):
    assert client.get("/api/sheets/v1/row/99999").status_code == 404


def test_v2_download_applies_edits(client):
    g = client.get("/api/sheets/v2").json()
    cols = g["columns"]
    rows = [dict(r) for r in g["rows"]]
    rows[0][cols[0]] = "EDITED VALUE"
    resp = client.post(
        "/api/sheets/v2/download",
        json={"columns": cols, "rows": rows, "filename": "out.xlsx", "sheet": "V2.1"},
    )
    assert resp.status_code == 200
    assert "out.xlsx" in resp.headers["content-disposition"]
    wb = load_workbook(io.BytesIO(resp.content))
    ws = wb.active
    header = [c.value for c in ws[1]]
    assert header == cols
    assert ws.cell(row=2, column=1).value == "EDITED VALUE"
