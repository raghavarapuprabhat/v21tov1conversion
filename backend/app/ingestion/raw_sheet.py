"""Raw sheet reader — returns the uploaded workbook grid *as is*.

Unlike `excel_loader` (which normalises into canonical `V1Field`/`V2Field`), this
module preserves every original column and the raw cell text. It powers the two
"sheet viewer" tabs (read-only V1, editable + downloadable V2.1). The header row
is located with the same best-match detection used by the canonical loader, then
*all* columns from that row are read verbatim.
"""
from __future__ import annotations

import io
import threading
from pathlib import Path
from typing import Any, Optional

from openpyxl import Workbook, load_workbook

from app.ingestion.excel_loader import (
    V1_HEADERS,
    V2_HEADERS,
    IngestionError,
    _detect_header_row,
    _MATCH_THRESHOLD,
)

ROW_KEY = "__row"  # source Excel row number, used to target edits


def _cell_text(v: Any) -> str:
    return "" if v is None else str(v)


def _open(path: str | Path, expected: set[str]):
    """Open the workbook and locate the header row, or raise IngestionError."""
    p = Path(path)
    if not p.exists():
        raise IngestionError(f"workbook not found: {p}")
    wb = load_workbook(p, data_only=True)
    ws = wb.active
    hrow, hits, _ = _detect_header_row(ws, expected)
    if hrow is None or hits < len(expected) * _MATCH_THRESHOLD:
        raise IngestionError(
            f"header row not found in {p.name} (matched {hits}/{len(expected)})"
        )
    return ws, hrow


def _columns(ws, hrow: int) -> list[str]:
    ncols = ws.max_column or 0
    columns: list[str] = []
    seen: dict[str, int] = {}
    for c in range(1, ncols + 1):
        raw = ws.cell(row=hrow, column=c).value
        label = str(raw).strip() if raw is not None else f"Column {c}"
        if not label:
            label = f"Column {c}"
        if label in seen:
            seen[label] += 1
            label = f"{label} ({seen[label]})"
        else:
            seen[label] = 0
        columns.append(label)
    return columns


def _build_grid(path: str | Path, expected: set[str]) -> dict:
    ws, hrow = _open(path, expected)
    columns = _columns(ws, hrow)
    ncols = len(columns)
    rows: list[dict] = []
    # iter_rows is far faster than per-cell .cell() random access on large sheets
    for r, values in enumerate(
        ws.iter_rows(min_row=hrow + 1, max_col=ncols, values_only=True),
        start=hrow + 1,
    ):
        cells: dict[str, Any] = {}
        blank = True
        for i in range(ncols):
            val = _cell_text(values[i] if i < len(values) else None)
            if val.strip():
                blank = False
            cells[columns[i]] = val
        if blank:
            continue
        cells[ROW_KEY] = r
        rows.append(cells)

    by_row = {row[ROW_KEY]: row for row in rows}
    return {"sheet": ws.title, "columns": columns, "rows": rows, "_by_row": by_row}


# Parsed-grid cache keyed by (resolved path, tag). openpyxl parsing dominates the
# request, so we parse a workbook once and reuse it for the full-sheet view *and*
# every single-row fetch. The (mtime_ns, size) guard auto-invalidates on re-upload.
_GRID_CACHE: dict[tuple[str, str], tuple[int, int, dict]] = {}
_CACHE_LOCK = threading.Lock()


def _cached_grid(path: str | Path, expected: set[str], tag: str) -> dict:
    p = Path(path)
    if not p.exists():
        raise IngestionError(f"workbook not found: {p}")
    st = p.stat()
    key = (str(p.resolve()), tag)
    with _CACHE_LOCK:
        hit = _GRID_CACHE.get(key)
        if hit and hit[0] == st.st_mtime_ns and hit[1] == st.st_size:
            return hit[2]
    grid = _build_grid(p, expected)            # parse outside the lock
    with _CACHE_LOCK:
        _GRID_CACHE[key] = (st.st_mtime_ns, st.st_size, grid)
    return grid


def _public_grid(grid: dict) -> dict:
    """Strip the internal row index before returning to the API."""
    return {"sheet": grid["sheet"], "columns": grid["columns"], "rows": grid["rows"]}


def read_v1_grid(path: str | Path) -> dict:
    return _public_grid(_cached_grid(path, V1_HEADERS, "v1"))


def read_v2_grid(path: str | Path) -> dict:
    return _public_grid(_cached_grid(path, V2_HEADERS, "v2"))


def _read_row(path: str | Path, expected: set[str], tag: str, row_no: int) -> Optional[dict]:
    """Return a single raw row (all columns, as is) by its source Excel row number.
    Served from the cached grid — O(1) after the first parse."""
    grid = _cached_grid(path, expected, tag)
    row = grid["_by_row"].get(row_no)
    if row is None:
        return None
    return {"sheet": grid["sheet"], "columns": grid["columns"], "row": row}


def read_v1_row(path: str | Path, row_no: int) -> Optional[dict]:
    return _read_row(path, V1_HEADERS, "v1", row_no)


def read_v2_row(path: str | Path, row_no: int) -> Optional[dict]:
    return _read_row(path, V2_HEADERS, "v2", row_no)


def write_grid_xlsx(columns: list[str], rows: list[dict], sheet_title: str = "Sheet1") -> bytes:
    """Serialise an (edited) grid back into an .xlsx workbook."""
    wb = Workbook()
    ws = wb.active
    ws.title = (sheet_title or "Sheet1")[:31]
    ws.append(columns)
    for row in rows:
        ws.append([_cell_text(row.get(c, "")) for c in columns])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
