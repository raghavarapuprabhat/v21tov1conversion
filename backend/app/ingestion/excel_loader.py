"""Excel loader + header detection + column mapping (LLD v1.2 §3.1).

Reads the V1 and V2.1 workbooks into canonical `V1Field` / `V2Field` records.
The header row is detected by best-match against the known header vocabulary
(the sample sheets carry a merged 'Table 1' title in row 1 and real headers in
row 2). If the match is below threshold the loader raises `IngestionError`
(surfaced as HTTP 422) rather than guessing.
"""
from __future__ import annotations

from pathlib import Path
from typing import Any, Callable, Optional

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from app.ingestion import normalize as N
from app.models.canonical import SourceRef, V1Field, V2Field

# --- known headers (the contract; values vary, headers do not) ----------------

V1_HEADERS = {
    "IS Reference Number", "CC DD Ref No", "Node",
    "Level 1", "Level 2", "Level 3", "Level 4",
    "Level 5", "Level 6", "Level 7", "Level 8",
    "Attribute", "XSD Field Type", "Nullable",
    "Min Occurrence", "Max Occurrence",
}

V2_HEADERS = {
    "Schema Name + JSON Path", "CLMT IS Reference Number", "Version Number",
    "Change Log", "Attribute CLM ID", "CCDM Attribute Name", "Source DD#",
    "CC_V1_Mapping Entity", "CC_V1_Mapping RP IND", "CC_V1_Mapping RP ORG",
    "Remarks For Repeating Block", "Node / Element", "Schema Name",
    "JSON Attribute Name", "Data Type", "Min Occurrence", "Max Occurrence",
    "Mandatory / Optional", "Schema + JSON Path + Attribute", "Mapping Remarks",
}

_MATCH_THRESHOLD = 0.6   # fraction of expected headers that must be present


class IngestionError(Exception):
    """Raised when a workbook cannot be parsed (e.g. header row not found)."""


# --- header detection ---------------------------------------------------------

def _detect_header_row(ws, expected: set[str], scan: int = 12):
    """Return (row_index, hits, {header: column_index}) for the best-matching row."""
    best = (None, 0, {})  # type: tuple[Optional[int], int, dict[str, int]]
    max_row = min(scan, ws.max_row or scan)
    for r in range(1, max_row + 1):
        colmap: dict[str, int] = {}
        for c in range(1, (ws.max_column or 0) + 1):
            label = N.clean(ws.cell(row=r, column=c).value)
            if label and label in expected and label not in colmap:
                colmap[label] = c
        if len(colmap) > best[1]:
            best = (r, len(colmap), colmap)
    return best


def _row_getter(ws, row: int, colmap: dict[str, int]) -> Callable[[str], Any]:
    def get(header: str) -> Any:
        col = colmap.get(header)
        return ws.cell(row=row, column=col).value if col else None
    return get


def _is_blank_row(get: Callable[[str], Any], headers: set[str]) -> bool:
    return all(N.clean(get(h)) is None for h in headers)


# --- V1 -----------------------------------------------------------------------

def load_v1(path: str | Path) -> list[V1Field]:
    p = Path(path)
    if not p.exists():
        raise IngestionError(f"V1 workbook not found: {p}")
    wb = load_workbook(p, data_only=True)
    ws = wb.active
    hrow, hits, colmap = _detect_header_row(ws, V1_HEADERS)
    if hrow is None or hits < len(V1_HEADERS) * _MATCH_THRESHOLD:
        raise IngestionError(
            f"V1 header row not found in {p.name} (matched {hits}/{len(V1_HEADERS)})"
        )

    fields: list[V1Field] = []
    for r in range(hrow + 1, (ws.max_row or hrow) + 1):
        get = _row_getter(ws, r, colmap)
        if _is_blank_row(get, V1_HEADERS):
            continue
        levels = [N.clean(get(f"Level {i}")) for i in range(1, 9)]
        while levels and levels[-1] is None:
            levels.pop()
        fields.append(V1Field(
            is_number_raw=N.clean(get("IS Reference Number")),
            is_number=N.norm_code(get("IS Reference Number")),
            dd_ref_raw=N.clean(get("CC DD Ref No")),
            dd_ref=N.norm_code(get("CC DD Ref No")),
            node_kind=N.clean(get("Node")),
            path=[lvl for lvl in levels if lvl is not None],
            attribute=N.clean(get("Attribute")),
            xsd_type_raw=N.clean(get("XSD Field Type")),
            xsd_type=N.norm_type_token(get("XSD Field Type")),
            nullable=N.norm_bool(get("Nullable")),
            min_occurs=N.parse_int(get("Min Occurrence")),
            max_occurs=N.parse_occurs(get("Max Occurrence")),
            source=SourceRef(sheet=ws.title, row=r),
        ))
    return fields


# --- V2.1 ---------------------------------------------------------------------

def load_v2(path: str | Path) -> list[V2Field]:
    p = Path(path)
    if not p.exists():
        raise IngestionError(f"V2.1 workbook not found: {p}")
    wb = load_workbook(p, data_only=True)
    ws = wb.active
    hrow, hits, colmap = _detect_header_row(ws, V2_HEADERS)
    if hrow is None or hits < len(V2_HEADERS) * _MATCH_THRESHOLD:
        raise IngestionError(
            f"V2.1 header row not found in {p.name} (matched {hits}/{len(V2_HEADERS)})"
        )

    fields: list[V2Field] = []
    for r in range(hrow + 1, (ws.max_row or hrow) + 1):
        get = _row_getter(ws, r, colmap)
        if _is_blank_row(get, V2_HEADERS):
            continue
        fields.append(V2Field(
            json_path=N.clean(get("Schema Name + JSON Path")),
            clmt_is_ref=N.clean(get("CLMT IS Reference Number")),
            version=N.clean(get("Version Number")),
            change_log=N.clean(get("Change Log")),
            clm_id=N.clean(get("Attribute CLM ID")),
            attribute_name=N.clean(get("CCDM Attribute Name")),
            dd_ref_raw=N.clean(get("Source DD#")),
            dd_ref=N.norm_code(get("Source DD#")),
            map_entity_raw=N.clean(get("CC_V1_Mapping Entity")),
            map_rp_ind_raw=N.clean(get("CC_V1_Mapping RP IND")),
            map_rp_org_raw=N.clean(get("CC_V1_Mapping RP ORG")),
            map_entity=N.norm_code(get("CC_V1_Mapping Entity")),
            map_rp_ind=N.norm_code(get("CC_V1_Mapping RP IND")),
            map_rp_org=N.norm_code(get("CC_V1_Mapping RP ORG")),
            repeat_remarks=N.clean(get("Remarks For Repeating Block")),
            node_kind=N.clean(get("Node / Element")),
            schema_name=N.clean(get("Schema Name")),
            json_attr=N.clean(get("JSON Attribute Name")),
            data_type_raw=N.clean(get("Data Type")),
            data_type=N.norm_type_token(get("Data Type")),
            min_occurs=N.parse_int(get("Min Occurrence")),
            max_occurs=N.parse_occurs(get("Max Occurrence")),
            mandatory_optional=N.clean(get("Mandatory / Optional")),
            full_path=N.clean(get("Schema + JSON Path + Attribute")),
            mapping_remarks=N.clean(get("Mapping Remarks")),
            source=SourceRef(sheet=ws.title, row=r),
        ))
    return fields


# Re-exported for callers that want a column letter for a header (audit/UI).
def column_letter(colmap: dict[str, int], header: str) -> Optional[str]:
    col = colmap.get(header)
    return get_column_letter(col) if col else None
